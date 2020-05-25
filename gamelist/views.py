from django.shortcuts import render
from .models import Game
import openpyxl

def home_page(request):
	return render(request, 'gamelist/home_page.html', {})

def game_list(request):
	all_games = Game.objects.order_by('game_id')
	all_age = set(Game.objects.all().values_list('age', flat=True))
	all_year = set(Game.objects.all().values_list('pub_year', flat=True))
	all_cat = set(Game.objects.all().values_list('category', flat=True))
	all_pub = set(Game.objects.all().values_list('publisher', flat=True))
	return render(request, 'gamelist/game_list.html', {'all_games': all_games, 'all_age': all_age, 'all_year': all_year, 'all_cat': all_cat, 'all_pub': all_pub})

def updatedb(request):
	workbook = openpyxl.load_workbook("D:/django/testdb.xlsx")
	sheet = workbook['games']
	y = 2
	all_games_id = Game.objects.all().values_list('game_id', flat=True)
	while y <= sheet.max_row:
		if sheet.cell(row = y, column = 1).value in all_games_id:
			y += 1
			continue
		else:
			Game.objects.create(game_id = sheet.cell(row = y, column = 1).value, title = sheet.cell(row = y, column = 2).value, description = sheet.cell(row = y, column = 3).value, min_players = sheet.cell(row = y, column = 4).value, max_players = sheet.cell(row = y, column = 5).value, age = sheet.cell(row = y, column = 6).value, pub_year = sheet.cell(row = y, column = 7).value, category = sheet.cell(row = y, column = 8).value, publisher = sheet.cell(row = y, column = 9).value,)
			y += 1
	return render(request, 'gamelist/updatedb.html', {})


